import streamlit as st
from openpyxl import load_workbook

# Excelファイル操作
def write_to_excel(file_path, sheet_name, cell, value):
    try:
        wb = load_workbook(filename=file_path, keep_vba=True)
        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(file_path)
    except Exception as e:
        raise RuntimeError(f"Excelファイルの書き込み中にエラーが発生しました: {e}")
    finally:
        wb.close()

def read_from_excel(file_path, sheet_name, cell):
    try:
        wb = load_workbook(filename=file_path, data_only=True, keep_vba=True)
        ws = wb[sheet_name]
        value = ws[cell].value
        return value
    except Exception as e:
        raise RuntimeError(f"Excelファイルの読み取り中にエラーが発生しました: {e}")
    finally:
        wb.close()

# Streamlitアプリ
def main():
    st.title("Webアプリ ⇔ Excel 連携")

    excel_path = "rennsyuu.xlsm"
    sheet = "Sheet1"

    # 項目と選択肢リスト
    categories = ["認知力・操作", "言語理解", "表出言語"]
    options = ["0〜3か月", "3〜6か月", "6〜9か月", "9〜12か月"]

    # 選択結果を格納する辞書
    selected_options = {}

    # 各項目ごとに選択肢を表示
    for index, category in enumerate(categories, start=1):
        st.subheader(category)
        selected_options[category] = st.radio(f"{category}の選択肢を選んでください:", options, key=f"radio_{index}")

    if st.button("Excelに書き込む"):
        try:
            # 各項目を対応するセルに書き込む（例: A3, B3, C3）
            for index, (category, selected_option) in enumerate(selected_options.items(), start=1):
                write_to_excel(excel_path, sheet, f"A{index + 2}", category)  # A3, A4, A5に項目
                write_to_excel(excel_path, sheet, f"B{index + 2}", selected_option)  # B3, B4, B5に選択肢
            st.success("各項目と選択肢がExcelに書き込まれました！")
        except RuntimeError as e:
            st.error(f"エラー: {e}")

    if st.button("Excelの答えを取得"):
        try:
            # A5セルの値を取得して表示
            result = read_from_excel(excel_path, sheet, "A5")
            st.write(f"Excelの答え: {result}")
        except RuntimeError as e:
            st.error(f"エラー: {e}")

if __name__ == "__main__":
    main()
